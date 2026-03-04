from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from med_exam_toolkit.models import Question
from med_exam_toolkit.exporters import register
from med_exam_toolkit.exporters.base import BaseExporter

# 中文表头映射
HEADER_LABELS = {
    "fingerprint":    "指纹",
    "pkg":            "来源",
    "cls":            "题库",
    "unit":           "章节",
    "mode":           "题型",
    "stem":           "共享题干",
    "sub_index":      "子题序号",
    "text":           "题目",
    "options":        "选项",
    "answer":         "答案",
    "answer_source":  "答案来源",
    "rate":           "正确率",
    "error_prone":    "易错项",
    "discuss":        "解析",
    "discuss_source": "解析来源",
    "point":          "考点",
    "ai_answer":      "AI答案",
    "ai_discuss":     "AI解析",
    "ai_confidence":  "AI置信度",
    "ai_model":       "AI模型",
}
for i in range(10):
    HEADER_LABELS[f"option_{chr(65 + i)}"] = f"选项{chr(65 + i)}"

COL_WIDTHS = {
    "text":           50,
    "discuss":        50,
    "ai_discuss":     50,
    "point":          40,
    "stem":           50,
    "unit":           20,
    "cls":            20,
    "options":        60,
    "answer_source":  12,
    "discuss_source": 12,
    "ai_answer":      12,
    "ai_confidence":  10,
    "ai_model":       16,
}
for i in range(10):
    COL_WIDTHS[f"option_{chr(65 + i)}"] = 25

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_AI_FILL     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_AI_COL_FILL = PatternFill(start_color="E8F4E8", end_color="E8F4E8", fill_type="solid")

# 这些列即使全空也强制保留（核心字段）
_ALWAYS_KEEP = {
    "text", "answer", "discuss", "mode", "unit", "cls",
    "sub_index", "options",
}

# AI 相关列（用绿色表头区分）
_AI_COLS = {"ai_answer", "ai_discuss", "ai_confidence", "ai_model"}


@register("xlsx")
class XlsxExporter(BaseExporter):

    def export(self, questions: list[Question], output_path: Path, **kwargs) -> None:
        split_options = kwargs.get("split_options", True)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        fp = output_path.with_suffix(".xlsx")

        rows, columns = self.flatten(questions, split_options=split_options)

        # ── 剔除全空列（_ALWAYS_KEEP 中的列强制保留）──
        active_columns = [
            col for col in columns
            if col in _ALWAYS_KEEP
            or any(row.get(col) not in (None, "", 0) for row in rows)
        ]
        hidden = len(columns) - len(active_columns)

        wb = Workbook()
        ws = wb.active
        ws.title = "题目"

        header_font    = Font(bold=True, color="FFFFFF")
        ai_header_font = Font(bold=True, color="1F6B2E")

        col_index = {key: i + 1 for i, key in enumerate(active_columns)}

        # 表头
        for col_idx, col_key in enumerate(active_columns, 1):
            cell = ws.cell(
                row=1, column=col_idx,
                value=HEADER_LABELS.get(col_key, col_key),
            )
            if col_key in _AI_COLS:
                cell.font = ai_header_font
                cell.fill = _AI_COL_FILL
            else:
                cell.font = header_font
                cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_key in enumerate(active_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_key, ""))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # 有效答案/解析来自 AI → 浅黄背景
            if row.get("answer_source") == "ai" and "answer" in col_index:
                ws.cell(row=row_idx, column=col_index["answer"]).fill = _AI_FILL
            if row.get("discuss_source") == "ai" and "discuss" in col_index:
                ws.cell(row=row_idx, column=col_index["discuss"]).fill = _AI_FILL

            # AI 原始列有内容 → 浅黄背景
            if row.get("ai_discuss") and "ai_discuss" in col_index:
                ws.cell(row=row_idx, column=col_index["ai_discuss"]).fill = _AI_FILL
            if row.get("ai_answer") and "ai_answer" in col_index:
                ws.cell(row=row_idx, column=col_index["ai_answer"]).fill = _AI_FILL

        # 列宽
        for col_idx, col_key in enumerate(active_columns, 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = COL_WIDTHS.get(col_key, 14)

        # 冻结首行
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(active_columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

        wb.save(fp)

        ai_cnt = sum(1 for r in rows if r.get("ai_discuss") or r.get("ai_answer"))
        hidden_note = f", 隐藏空列: {hidden}" if hidden else ""
        print(f"[INFO] XLSX 导出完成: {fp} ({len(rows)} 行, {len(active_columns)} 列{hidden_note}, 含AI内容: {ai_cnt} 行)")