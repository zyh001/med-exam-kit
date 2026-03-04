import json
import tempfile
from pathlib import Path
from med_exam_toolkit.loader import load_json_files
from med_exam_toolkit.dedup import deduplicate, compute_fingerprint
from med_exam_toolkit.exporters import discover as discover_exporters, get_exporter
from med_exam_toolkit.exporters.base import BaseExporter

# 3 个选项
Q_3_OPTIONS = {
    "name": "t1", "cls": "测试", "pkg": "ahuyikao.com",
    "numb": "1/1", "unit": "测试章节", "mode": "A1型题",
    "test": "以下哪项正确",
    "option": ["A.选项1", "B.选项2", "C.选项3"],
    "answer": "A", "rate": "80%", "error_prone": "B", "discuss": "解析"
}

# 5 个选项（标准）
Q_5_OPTIONS = {
    "name": "t2", "cls": "测试", "pkg": "ahuyikao.com",
    "numb": "2/2", "unit": "测试章节", "mode": "A1型题",
    "test": "标准五选一",
    "option": ["A.甲", "B.乙", "C.丙", "D.丁", "E.戊"],
    "answer": "C", "rate": "60%", "error_prone": "A", "discuss": "解析"
}

# 8 个选项
Q_8_OPTIONS = {
    "name": "t3", "cls": "测试", "pkg": "ahuyikao.com",
    "numb": "3/3", "unit": "测试章节", "mode": "A1型题",
    "test": "八选一题目",
    "option": [f"{chr(65+i)}.选项{i+1}" for i in range(8)],
    "answer": "F", "rate": "30%", "error_prone": "A", "discuss": "解析"
}

# 10 个选项
Q_10_OPTIONS = {
    "name": "t4", "cls": "测试", "pkg": "ahuyikao.com",
    "numb": "4/4", "unit": "测试章节", "mode": "A1型题",
    "test": "十选一题目",
    "option": [f"{chr(65+i)}.选项{i+1}" for i in range(10)],
    "answer": "J", "rate": "15%", "error_prone": "A", "discuss": "解析"
}

ALL_SAMPLES = [Q_3_OPTIONS, Q_5_OPTIONS, Q_8_OPTIONS, Q_10_OPTIONS]


def _setup(tmpdir: Path):
    for i, s in enumerate(ALL_SAMPLES):
        (tmpdir / f"q_{i}.json").write_text(json.dumps(s, ensure_ascii=False), encoding="utf-8")
    return load_json_files(str(tmpdir), {"ahuyikao.com": "ahuyikao"})


def test_parse_variable_options():
    """验证不同选项数量都能正确解析"""
    with tempfile.TemporaryDirectory() as tmpdir:
        questions = _setup(Path(tmpdir))
        assert len(questions) == 4

        option_counts = [len(q.sub_questions[0].options) for q in questions]
        assert sorted(option_counts) == [3, 5, 8, 10]


def test_flatten_split_mode():
    """拆分模式：动态列数 = 最大选项数"""
    with tempfile.TemporaryDirectory() as tmpdir:
        questions = _setup(Path(tmpdir))
        rows, columns = BaseExporter.flatten(questions, split_options=True)

        # 最大 10 个选项 → 应有 option_A 到 option_J
        opt_cols = [c for c in columns if c.startswith("option_")]
        assert len(opt_cols) == 10

        # 3 选项的题，option_D ~ option_J 应为空
        q3_row = [r for r in rows if r["text"] == "以下哪项正确"][0]
        assert q3_row["option_A"] == "A.选项1"
        assert q3_row["option_C"] == "C.选项3"
        assert q3_row["option_D"] == ""
        assert q3_row["option_J"] == ""

        # 10 选项的题，所有列都有值
        q10_row = [r for r in rows if r["text"] == "十选一题目"][0]
        assert q10_row["option_J"] == "J.选项10"


def test_flatten_merge_mode():
    """合并模式：所有选项拼成一个字符串"""
    with tempfile.TemporaryDirectory() as tmpdir:
        questions = _setup(Path(tmpdir))
        rows, columns = BaseExporter.flatten(questions, split_options=False)

        assert "options" in columns
        assert "option_A" not in columns

        q8_row = [r for r in rows if r["text"] == "八选一题目"][0]
        assert q8_row["options"].count("|") == 7  # 8 个选项，7 个分隔符


def test_fingerprint_different_option_counts():
    """不同选项数量的题目指纹应不同"""
    with tempfile.TemporaryDirectory() as tmpdir:
        questions = _setup(Path(tmpdir))
        fps = [compute_fingerprint(q) for q in questions]
        assert len(set(fps)) == 4  # 全部不同


def test_xlsx_export_variable_options():
    """XLSX 导出能正确处理不同选项数"""
    discover_exporters()
    with tempfile.TemporaryDirectory() as tmpdir:
        questions = _setup(Path(tmpdir))
        out = Path(tmpdir) / "output" / "test"
        get_exporter("xlsx").export(questions, out, split_options=True)
        fp = out.with_suffix(".xlsx")
        assert fp.exists()

        # 验证文件可以被 openpyxl 读回
        from openpyxl import load_workbook
        wb = load_workbook(fp)
        ws = wb.active

        # 表头行应包含 option_A 到 option_J
        headers = [cell.value for cell in ws[1]]
        assert "选项A" in headers
        assert "选项J" in headers

        # 数据行数 = 4 题 + 1 表头
        assert ws.max_row == 5


def test_csv_export_variable_options():
    """CSV 导出两种模式都能工作"""
    discover_exporters()
    with tempfile.TemporaryDirectory() as tmpdir:
        questions = _setup(Path(tmpdir))

        # 拆分模式
        out1 = Path(tmpdir) / "output" / "split"
        get_exporter("csv").export(questions, out1, split_options=True)
        content1 = out1.with_suffix(".csv").read_text(encoding="utf-8-sig")
        assert "option_A" in content1

        # 合并模式
        out2 = Path(tmpdir) / "output" / "merge"
        get_exporter("csv").export(questions, out2, split_options=False)
        content2 = out2.with_suffix(".csv").read_text(encoding="utf-8-sig")
        assert "options" in content2
        assert "option_A" not in content2


if __name__ == "__main__":
    test_parse_variable_options()
    test_flatten_split_mode()
    test_flatten_merge_mode()
    test_fingerprint_different_option_counts()
    test_xlsx_export_variable_options()
    test_csv_export_variable_options()
    print("All variable options tests passed!")
